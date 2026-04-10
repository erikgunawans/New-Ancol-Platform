"""Excel report generation using openpyxl.

Produces pivot-ready compliance findings with severity color coding.
"""

from __future__ import annotations

import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

SEVERITY_COLORS = {
    "critical": "FF1744",  # Red
    "high": "FF6D00",  # Orange
    "medium": "FFAB00",  # Amber
    "low": "00C853",  # Green
}

HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def generate_excel(
    document_id: str,
    meeting_date: str,
    meeting_number: str,
    scorecard: dict,
    findings: list[dict],
    corrective_suggestions: list[dict],
) -> bytes:
    """Generate Excel report with multiple sheets."""
    wb = Workbook()

    # Sheet 1: Scorecard
    _create_scorecard_sheet(wb, scorecard, meeting_number, meeting_date)

    # Sheet 2: Findings
    _create_findings_sheet(wb, findings)

    # Sheet 3: Corrective Suggestions
    _create_corrections_sheet(wb, corrective_suggestions)

    # Remove default sheet if still present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_scorecard_sheet(wb: Workbook, scorecard: dict, meeting_number: str, meeting_date: str):
    ws = wb.create_sheet("Scorecard", 0)
    ws.sheet_properties.tabColor = "1A237E"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "SCORECARD KEPATUHAN RISALAH RAPAT"
    ws["A1"].font = Font(size=14, bold=True, color="1A237E")

    ws["A3"] = "Nomor Rapat"
    ws["B3"] = meeting_number
    ws["A4"] = "Tanggal Rapat"
    ws["B4"] = meeting_date

    # Scores
    headers = ["Pilar", "Skor", "Bobot", "Kontribusi"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    pillars = [
        ("Struktural", scorecard.get("structural_score", 0), 0.30),
        ("Substantif", scorecard.get("substantive_score", 0), 0.35),
        ("Regulasi", scorecard.get("regulatory_score", 0), 0.35),
    ]

    for i, (name, score, weight) in enumerate(pillars, 7):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=round(score, 1))
        ws.cell(row=i, column=3, value=f"{weight:.0%}")
        ws.cell(row=i, column=4, value=round(score * weight, 1))

    ws.cell(row=10, column=1, value="KOMPOSIT").font = Font(bold=True)
    ws.cell(row=10, column=2, value=scorecard.get("composite_score", 0)).font = Font(
        bold=True, size=14
    )

    # Column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_findings_sheet(wb: Workbook, findings: list[dict]):
    ws = wb.create_sheet("Temuan")
    ws.sheet_properties.tabColor = "FF6D00"

    headers = [
        "No",
        "Severity",
        "Keputusan",
        "Regulasi",
        "Status",
        "Judul",
        "Deskripsi",
        "Red Flag",
        "Tipe Red Flag",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, finding in enumerate(
        sorted(findings, key=lambda f: _severity_order(f.get("severity", "low"))),
        start=2,
    ):
        severity = finding.get("severity", "medium")
        color = SEVERITY_COLORS.get(severity, "FFFFFF")

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        severity_cell = ws.cell(row=row_idx, column=2, value=severity.upper())
        severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        severity_cell.font = Font(color="FFFFFF", bold=True)

        ws.cell(row=row_idx, column=3, value=finding.get("resolution_number", ""))
        ws.cell(row=row_idx, column=4, value=finding.get("regulation_id", ""))
        ws.cell(row=row_idx, column=5, value=finding.get("compliance_status", ""))
        ws.cell(row=row_idx, column=6, value=finding.get("title", ""))
        ws.cell(row=row_idx, column=7, value=finding.get("description", ""))
        ws.cell(row=row_idx, column=8, value="Ya" if finding.get("is_red_flag") else "Tidak")
        ws.cell(row=row_idx, column=9, value=finding.get("red_flag_type", ""))

    # Auto-width
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Auto-filter
    ws.auto_filter.ref = f"A1:I{len(findings) + 1}"


def _create_corrections_sheet(wb: Workbook, corrections: list[dict]):
    ws = wb.create_sheet("Saran Perbaikan")
    ws.sheet_properties.tabColor = "00C853"

    headers = [
        "Finding ID",
        "Permasalahan",
        "Redaksi Saat Ini",
        "Saran Perbaikan",
        "Dasar Regulasi",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, cs in enumerate(corrections, start=2):
        ws.cell(row=row_idx, column=1, value=cs.get("finding_id", ""))
        ws.cell(row=row_idx, column=2, value=cs.get("issue_explanation", ""))
        ws.cell(row=row_idx, column=3, value=cs.get("current_wording", ""))
        ws.cell(row=row_idx, column=4, value=cs.get("suggested_wording", ""))
        ws.cell(row=row_idx, column=5, value=cs.get("regulatory_basis", ""))

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 30


def _severity_order(severity: str) -> int:
    return {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(severity, 4)
